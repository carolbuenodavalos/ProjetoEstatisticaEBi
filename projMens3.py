from __future__ import annotations

import json
import math
import re
import shutil
from datetime import date
from pathlib import Path

try:
    import numpy as np
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
    from openpyxl.utils import get_column_letter
    from PIL import Image, ImageDraw, ImageFont
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Image as RLImage,
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )
except ModuleNotFoundError as exc:
    missing = exc.name
    raise SystemExit(
        f"Dependência não instalada: {missing}\n"
        "Execute primeiro:\n"
        "  python -m pip install -r requirements.txt\n"
        "Depois rode novamente:\n"
        "  python projMens3.py"
    ) from exc


BASE_DIR = Path(__file__).resolve().parent
RAW_PATH = BASE_DIR / "dados_brutos" / "dadosSIM2005-.csv"
TREATED_DIR = BASE_DIR / "dados_tratados"
DOC_DIR = BASE_DIR / "documentacao"
NOTEBOOK_DIR = BASE_DIR / "notebooks"
GRAPH_DIR = DOC_DIR / "graficos"
DELIVERY_DIR = BASE_DIR / "PM3_Tratamento_Dados"

ACCESS_DATE = "16/05/2026"
SOURCE_URL = "https://dadosabertos.saude.gov.br/dataset/sim"
SOURCE_DESC = (
    "Sistema de Informação sobre Mortalidade (SIM), Ministério da Saúde/DATASUS, "
    "com registros de óbitos e variáveis socioeconômicas, local de residência, "
    "local de ocorrência e causa básica do óbito."
)


RACA_MAP = {
    1: "Branca",
    2: "Preta",
    3: "Amarela",
    4: "Parda",
    5: "Indígena",
    9: "Ignorada",
}

SEXO_MAP = {
    1: "Masculino",
    2: "Feminino",
    0: "Ignorado",
    9: "Ignorado",
}

LOCAL_MAP = {
    1: "Hospital",
    2: "Outro estabelecimento de saúde",
    3: "Domicílio",
    4: "Via pública",
    5: "Outros",
    9: "Ignorado",
}


def ensure_dirs() -> None:
    for folder in [TREATED_DIR, DOC_DIR, NOTEBOOK_DIR, GRAPH_DIR]:
        folder.mkdir(parents=True, exist_ok=True)


def normalize_column_name(name: str) -> str:
    name = name.strip().lower()
    name = re.sub(r"[^a-z0-9_]+", "_", name)
    name = re.sub(r"_+", "_", name)
    return name.strip("_")


def read_raw() -> pd.DataFrame:
    df = pd.read_csv(RAW_PATH)
    df.columns = [normalize_column_name(c) for c in df.columns]
    return df


def cause_group(code: str) -> str:
    code = str(code).strip().upper()
    if not code:
        return "Não informado"
    first = code[0]
    if first in {"A", "B"}:
        return "Doenças infecciosas e parasitárias"
    if first in {"C", "D"}:
        return "Neoplasias e sangue"
    if first == "E":
        return "Endócrinas, nutricionais e metabólicas"
    if first == "G":
        return "Sistema nervoso"
    if first == "I":
        return "Aparelho circulatório"
    if first == "J":
        return "Aparelho respiratório"
    if first == "P":
        return "Afecções perinatais"
    if first == "Q":
        return "Malformações congênitas"
    if first == "R":
        return "Sintomas e causas mal definidas"
    if first in {"V", "W", "X", "Y"}:
        return "Causas externas"
    return "Outras causas"


def avoidability(code: str) -> str:
    group = cause_group(code)
    if group == "Causas externas":
        return "Evitável - causas externas"
    if group in {
        "Doenças infecciosas e parasitárias",
        "Aparelho respiratório",
        "Endócrinas, nutricionais e metabólicas",
        "Afecções perinatais",
    }:
        return "Potencialmente evitável por atenção à saúde"
    if group == "Sintomas e causas mal definidas":
        return "Indefinida - revisar qualidade do registro"
    return "Não classificada como evitável"


def age_band(age_years: float) -> str:
    if age_years < 1 / 12:
        return "Neonatal"
    if age_years < 1:
        return "Pós-neonatal"
    if age_years < 2:
        return "1 ano"
    if age_years < 3:
        return "2 anos"
    if age_years < 4:
        return "3 anos"
    return "4 anos"


def age_group_bi(age_years: float) -> str:
    if age_years < 1:
        return "Menor de 1 ano"
    return "1 a 4 anos"


def quality_report_before(df: pd.DataFrame) -> pd.DataFrame:
    missing = df.isna().sum()
    problems = [
        {
            "problema": "Valores ausentes",
            "colunas_afetadas": ", ".join(missing[missing > 0].index.tolist()) or "Nenhuma",
            "quantidade": int(missing.sum()),
            "decisao": "Preencher raça/cor como Ignorada; remover escolaridade por excesso de ausência.",
        },
        {
            "problema": "Linhas duplicadas",
            "colunas_afetadas": "Todas",
            "quantidade": int(df.duplicated().sum()),
            "decisao": "Remover duplicidades exatas para evitar dupla contagem.",
        },
        {
            "problema": "Coluna com baixa utilidade analítica",
            "colunas_afetadas": "escolaridade",
            "quantidade": int(df["escolaridade"].isna().sum()) if "escolaridade" in df.columns else 0,
            "decisao": "Remover do dataset final, pois mais de 90% dos registros estão vazios.",
        },
        {
            "problema": "Categorias codificadas",
            "colunas_afetadas": "sexo, raca_cor, local_ocorrencia",
            "quantidade": 3,
            "decisao": "Criar descrições padronizadas para uso em BI e análise exploratória.",
        },
        {
            "problema": "Outliers estatísticos",
            "colunas_afetadas": "idade",
            "quantidade": 0,
            "decisao": "Calcular pelo método IQR após remoção de duplicidades e manter com coluna indicadora.",
        },
    ]
    return pd.DataFrame(problems)


def treat_data(raw: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    df = raw.copy()
    initial_rows = len(df)
    initial_cols = df.shape[1]
    missing_before = df.isna().sum().rename("faltantes_antes").reset_index()
    missing_before = missing_before.rename(columns={"index": "coluna"})

    df = df.drop_duplicates().copy()
    duplicates_removed = initial_rows - len(df)

    escolaridade_missing_pct = df["escolaridade"].isna().mean() * 100
    df = df.drop(columns=["escolaridade"])

    df["causa_basica"] = df["causa_basica"].astype(str).str.strip().str.upper()
    df["sigla_uf"] = df["sigla_uf"].astype(str).str.strip().str.upper()
    df["raca_cor"] = df["raca_cor"].fillna(9).astype(int)
    df["sexo"] = df["sexo"].astype(int)
    df["local_ocorrencia"] = df["local_ocorrencia"].astype(int)
    df["ano"] = df["ano"].astype(int)
    df["id_municipio_residencia"] = df["id_municipio_residencia"].astype(int)
    df["idade"] = pd.to_numeric(df["idade"], errors="coerce")
    df = df[df["idade"].between(0, 4, inclusive="both")].copy()

    df["sexo_desc"] = df["sexo"].map(SEXO_MAP).fillna("Ignorado")
    df["raca_cor_desc"] = df["raca_cor"].map(RACA_MAP).fillna("Ignorada")
    df["local_ocorrencia_desc"] = df["local_ocorrencia"].map(LOCAL_MAP).fillna("Ignorado")
    df["grupo_causa"] = df["causa_basica"].apply(cause_group)
    df["classificacao_evitabilidade"] = df["causa_basica"].apply(avoidability)
    df["idade_meses"] = (df["idade"] * 12).round(2)
    df["faixa_etaria"] = df["idade"].apply(age_band)
    df["faixa_etaria_bi"] = df["idade"].apply(age_group_bi)
    df["causa_externa"] = df["grupo_causa"].eq("Causas externas")
    df["obito_hospitalar"] = df["local_ocorrencia"].eq(1)

    q1 = df["idade"].quantile(0.25)
    q3 = df["idade"].quantile(0.75)
    iqr = q3 - q1
    lower = max(0, q1 - 1.5 * iqr)
    upper = q3 + 1.5 * iqr
    df["idade_outlier_iqr"] = ~df["idade"].between(lower, upper, inclusive="both")

    min_age = df["idade"].min()
    max_age = df["idade"].max()
    df["idade_normalizada_minmax"] = ((df["idade"] - min_age) / (max_age - min_age)).round(6)
    std = df["idade"].std(ddof=0)
    df["idade_padronizada_zscore"] = ((df["idade"] - df["idade"].mean()) / std).round(6)

    selected_columns = [
        "ano",
        "sigla_uf",
        "id_municipio_residencia",
        "causa_basica",
        "grupo_causa",
        "classificacao_evitabilidade",
        "idade",
        "idade_meses",
        "faixa_etaria",
        "faixa_etaria_bi",
        "sexo",
        "sexo_desc",
        "raca_cor",
        "raca_cor_desc",
        "local_ocorrencia",
        "local_ocorrencia_desc",
        "causa_externa",
        "obito_hospitalar",
        "idade_outlier_iqr",
        "idade_normalizada_minmax",
        "idade_padronizada_zscore",
    ]
    df = df[selected_columns].sort_values(
        ["ano", "causa_basica", "idade", "sexo_desc"], ignore_index=True
    )

    without_outliers = df[~df["idade_outlier_iqr"]].copy()
    missing_after = df.isna().sum().rename("faltantes_depois").reset_index()
    missing_after = missing_after.rename(columns={"index": "coluna"})
    missing_compare = missing_before.merge(missing_after, on="coluna", how="outer").fillna(0)
    missing_compare[["faltantes_antes", "faltantes_depois"]] = missing_compare[
        ["faltantes_antes", "faltantes_depois"]
    ].astype(int)

    summary = {
        "linhas_base_original": int(initial_rows),
        "colunas_base_original": int(initial_cols),
        "linhas_apos_tratamento": int(len(df)),
        "colunas_apos_tratamento": int(df.shape[1]),
        "duplicatas_removidas": int(duplicates_removed),
        "faltantes_antes": int(raw.isna().sum().sum()),
        "faltantes_depois": int(df.isna().sum().sum()),
        "escolaridade_faltante_pct": round(float(escolaridade_missing_pct), 2),
        "outliers_idade_iqr": int(df["idade_outlier_iqr"].sum()),
        "limite_iqr_idade_inferior": round(float(lower), 4),
        "limite_iqr_idade_superior": round(float(upper), 4),
        "taxa_retencao_pct": round(len(df) / initial_rows * 100, 2),
        "fonte": SOURCE_URL,
        "data_acesso": ACCESS_DATE,
    }
    return df, without_outliers, {"resumo": summary, "faltantes": missing_compare}


def save_tables(raw: pd.DataFrame, final: pd.DataFrame, no_outliers: pd.DataFrame, meta: dict) -> dict:
    final.to_csv(TREATED_DIR / "dataset_final_tratado.csv", sep=";", index=False, encoding="utf-8-sig")
    no_outliers.to_csv(
        TREATED_DIR / "dataset_final_tratado_sem_outliers.csv",
        sep=";",
        index=False,
        encoding="utf-8-sig",
    )

    quality = quality_report_before(raw)
    quality.loc[quality["problema"].eq("Outliers estatísticos"), "quantidade"] = meta["resumo"][
        "outliers_idade_iqr"
    ]
    quality.to_csv(DOC_DIR / "problemas_qualidade.csv", sep=";", index=False, encoding="utf-8-sig")
    meta["faltantes"].to_csv(DOC_DIR / "valores_faltantes_antes_depois.csv", sep=";", index=False, encoding="utf-8-sig")

    stats = final[["ano", "idade", "idade_meses", "idade_normalizada_minmax", "idade_padronizada_zscore"]].describe().round(4)
    stats.to_csv(DOC_DIR / "estatisticas_descritivas.csv", sep=";", encoding="utf-8-sig")

    cat_freq_frames = []
    for col in ["sexo_desc", "raca_cor_desc", "local_ocorrencia_desc", "grupo_causa", "faixa_etaria"]:
        tmp = final[col].value_counts().rename_axis("categoria").reset_index(name="quantidade")
        tmp.insert(0, "variavel", col)
        cat_freq_frames.append(tmp)
    categorical_freq = pd.concat(cat_freq_frames, ignore_index=True)
    categorical_freq.to_csv(DOC_DIR / "frequencias_categoricas.csv", sep=";", index=False, encoding="utf-8-sig")

    agg_year = final.groupby("ano", as_index=False).size().rename(columns={"size": "total_obitos"})
    agg_year.to_csv(DOC_DIR / "agregacao_obitos_por_ano.csv", sep=";", index=False, encoding="utf-8-sig")

    agg_year_band = final.groupby(["ano", "faixa_etaria_bi"], as_index=False).size().rename(columns={"size": "total_obitos"})
    agg_year_band.to_csv(DOC_DIR / "agregacao_obitos_por_ano_faixa.csv", sep=";", index=False, encoding="utf-8-sig")

    agg_cause = (
        final.groupby(["grupo_causa", "local_ocorrencia_desc"], as_index=False)
        .size()
        .rename(columns={"size": "total_obitos"})
        .sort_values("total_obitos", ascending=False)
    )
    agg_cause.to_csv(DOC_DIR / "agregacao_causa_local.csv", sep=";", index=False, encoding="utf-8-sig")

    cause_age_counts = (
        pd.crosstab(final["grupo_causa"], final["faixa_etaria_bi"])
        .reset_index()
        .rename_axis(None, axis=1)
    )
    age_cols = [c for c in cause_age_counts.columns if c != "grupo_causa"]
    cause_age_counts["total_obitos"] = cause_age_counts[age_cols].sum(axis=1)
    cause_age_counts = cause_age_counts.sort_values("total_obitos", ascending=False)
    cause_age_counts.to_csv(DOC_DIR / "cruzamento_causa_faixa_etaria.csv", sep=";", index=False, encoding="utf-8-sig")

    cause_age_pct = (
        pd.crosstab(final["grupo_causa"], final["faixa_etaria_bi"], normalize="index")
        .mul(100)
        .round(1)
        .reset_index()
        .rename_axis(None, axis=1)
    )
    cause_age_pct.to_csv(DOC_DIR / "cruzamento_causa_faixa_etaria_percentual.csv", sep=";", index=False, encoding="utf-8-sig")

    external_by_age = (
        final.groupby("faixa_etaria_bi")
        .agg(total_obitos=("causa_externa", "size"), causas_externas=("causa_externa", "sum"))
        .reset_index()
    )
    external_by_age["percentual_causas_externas"] = (
        external_by_age["causas_externas"] / external_by_age["total_obitos"] * 100
    ).round(1)
    external_by_age.to_csv(DOC_DIR / "cruzamento_causa_externa_faixa_etaria.csv", sep=";", index=False, encoding="utf-8-sig")

    local_summary = (
        final.groupby("grupo_causa")
        .agg(
            total_obitos=("grupo_causa", "size"),
            obitos_hospitalares=("obito_hospitalar", "sum"),
            obitos_domicilio=("local_ocorrencia_desc", lambda s: int((s == "Domicílio").sum())),
        )
        .reset_index()
    )
    local_summary["percentual_hospitalar"] = (
        local_summary["obitos_hospitalares"] / local_summary["total_obitos"] * 100
    ).round(1)
    local_summary["percentual_domicilio"] = (
        local_summary["obitos_domicilio"] / local_summary["total_obitos"] * 100
    ).round(1)
    local_summary = local_summary.sort_values("total_obitos", ascending=False)
    local_summary.to_csv(DOC_DIR / "cruzamento_causa_local_resumo.csv", sep=";", index=False, encoding="utf-8-sig")

    with open(DOC_DIR / "resumo_processamento.json", "w", encoding="utf-8") as f:
        json.dump(meta["resumo"], f, ensure_ascii=False, indent=2)

    return {
        "quality": quality,
        "stats": stats.reset_index().rename(columns={"index": "estatistica"}),
        "categorical_freq": categorical_freq,
        "agg_year": agg_year,
        "agg_year_band": agg_year_band,
        "agg_cause": agg_cause,
        "cause_age_counts": cause_age_counts,
        "cause_age_pct": cause_age_pct,
        "external_by_age": external_by_age,
        "local_summary": local_summary,
    }


def get_font(size: int = 13) -> ImageFont.ImageFont:
    candidates = [
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/calibri.ttf",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size)
    return ImageFont.load_default()


def draw_bar_chart(data: pd.Series, title: str, path: Path, color: tuple[int, int, int] = (47, 103, 246)) -> None:
    width, height = 1200, 720
    margin_l, margin_r, margin_t, margin_b = 110, 50, 80, 140
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    font = get_font(18)
    small = get_font(14)
    title_font = get_font(28)
    draw.text((margin_l, 25), title, fill=(30, 40, 55), font=title_font)
    values = data.values.astype(float)
    labels = [str(x) for x in data.index]
    max_value = max(values.max(), 1)
    chart_w = width - margin_l - margin_r
    chart_h = height - margin_t - margin_b
    n = len(values)
    gap = max(4, int(chart_w / max(n, 1) * 0.18))
    bar_w = max(8, int((chart_w - gap * (n - 1)) / max(n, 1)))
    draw.line((margin_l, margin_t, margin_l, margin_t + chart_h), fill=(120, 130, 145), width=2)
    draw.line((margin_l, margin_t + chart_h, width - margin_r, margin_t + chart_h), fill=(120, 130, 145), width=2)
    for i, value in enumerate(values):
        x0 = margin_l + i * (bar_w + gap)
        x1 = x0 + bar_w
        y1 = margin_t + chart_h
        y0 = y1 - int((value / max_value) * (chart_h - 20))
        draw.rectangle((x0, y0, x1, y1), fill=color)
        draw.text((x0, y0 - 22), str(int(value)), fill=(30, 40, 55), font=small)
        label = labels[i][:18]
        draw.text((x0, y1 + 10), label, fill=(30, 40, 55), font=small)
    img.save(path)


def draw_horizontal_bar_chart(
    data: pd.Series,
    title: str,
    path: Path,
    color: tuple[int, int, int] = (111, 90, 169),
) -> None:
    width, height = 1200, 720
    margin_l, margin_r, margin_t, margin_b = 360, 80, 80, 60
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    font = get_font(16)
    title_font = get_font(28)
    draw.text((70, 25), title, fill=(30, 40, 55), font=title_font)
    values = data.values.astype(float)
    labels = [str(x) for x in data.index]
    max_value = max(values.max(), 1)
    chart_w = width - margin_l - margin_r
    chart_h = height - margin_t - margin_b
    n = len(values)
    gap = 14
    bar_h = max(28, int((chart_h - gap * (n - 1)) / max(n, 1)))
    draw.line((margin_l, margin_t, margin_l, margin_t + chart_h), fill=(120, 130, 145), width=2)
    for i, value in enumerate(values):
        y0 = margin_t + i * (bar_h + gap)
        y1 = y0 + bar_h
        x1 = margin_l + int((value / max_value) * chart_w)
        label = labels[i]
        if len(label) > 34:
            label = label[:31] + "..."
        draw.text((70, y0 + max(0, (bar_h - 16) // 2)), label, fill=(30, 40, 55), font=font)
        draw.rectangle((margin_l, y0, x1, y1), fill=color)
        draw.text((x1 + 8, y0 + max(0, (bar_h - 16) // 2)), str(int(value)), fill=(30, 40, 55), font=font)
    img.save(path)


def draw_stacked_horizontal_chart(table: pd.DataFrame, title: str, path: Path) -> None:
    width, height = 1200, 760
    margin_l, margin_r, margin_t, margin_b = 360, 80, 110, 80
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    font = get_font(15)
    title_font = get_font(27)
    draw.text((70, 25), title, fill=(30, 40, 55), font=title_font)

    ordered = table.head(8).copy()
    columns = [c for c in ["Menor de 1 ano", "1 a 4 anos"] if c in ordered.columns]
    colors_by_col = {
        "Menor de 1 ano": (62, 146, 131),
        "1 a 4 anos": (216, 77, 68),
    }
    totals = ordered[columns].sum(axis=1)
    max_total = max(float(totals.max()), 1)
    chart_w = width - margin_l - margin_r
    chart_h = height - margin_t - margin_b
    n = len(ordered)
    gap = 14
    bar_h = max(30, int((chart_h - gap * (n - 1)) / max(n, 1)))

    legend_x = margin_l
    for col in columns:
        draw.rectangle((legend_x, 72, legend_x + 20, 92), fill=colors_by_col[col])
        draw.text((legend_x + 28, 70), col, fill=(30, 40, 55), font=font)
        legend_x += 190

    for i, row in enumerate(ordered.itertuples(index=False)):
        label = str(getattr(row, "grupo_causa"))
        if len(label) > 34:
            label = label[:31] + "..."
        y0 = margin_t + i * (bar_h + gap)
        y1 = y0 + bar_h
        draw.text((70, y0 + max(0, (bar_h - 15) // 2)), label, fill=(30, 40, 55), font=font)
        x = margin_l
        for col in columns:
            value = float(ordered.iloc[i][col])
            segment_w = int((value / max_total) * chart_w)
            if segment_w > 0:
                draw.rectangle((x, y0, x + segment_w, y1), fill=colors_by_col[col])
                if segment_w > 34:
                    draw.text((x + 6, y0 + max(0, (bar_h - 15) // 2)), str(int(value)), fill="white", font=font)
            x += segment_w
        draw.text((x + 8, y0 + max(0, (bar_h - 15) // 2)), str(int(totals.iloc[i])), fill=(30, 40, 55), font=font)
    img.save(path)


def draw_percent_bar_chart(data: pd.Series, title: str, path: Path) -> None:
    width, height = 950, 560
    margin_l, margin_r, margin_t, margin_b = 120, 60, 90, 90
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    font = get_font(18)
    title_font = get_font(26)
    draw.text((margin_l, 30), title, fill=(30, 40, 55), font=title_font)
    values = data.values.astype(float)
    labels = [str(x) for x in data.index]
    chart_w = width - margin_l - margin_r
    chart_h = height - margin_t - margin_b
    max_value = max(values.max(), 1)
    n = len(values)
    gap = 80 if n <= 3 else 30
    bar_w = max(70, int((chart_w - gap * (n - 1)) / max(n, 1)))
    draw.line((margin_l, margin_t + chart_h, width - margin_r, margin_t + chart_h), fill=(120, 130, 145), width=2)
    for i, value in enumerate(values):
        x0 = margin_l + i * (bar_w + gap)
        y1 = margin_t + chart_h
        y0 = y1 - int((value / max_value) * (chart_h - 20))
        draw.rectangle((x0, y0, x0 + bar_w, y1), fill=(216, 77, 68))
        draw.text((x0 + 8, y0 - 28), f"{value:.1f}%", fill=(30, 40, 55), font=font)
        draw.text((x0, y1 + 14), labels[i], fill=(30, 40, 55), font=font)
    img.save(path)


def draw_line_chart(data: pd.Series, title: str, path: Path) -> None:
    width, height = 1200, 700
    margin_l, margin_r, margin_t, margin_b = 100, 60, 80, 100
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    font = get_font(16)
    title_font = get_font(28)
    draw.text((margin_l, 25), title, fill=(30, 40, 55), font=title_font)
    values = data.values.astype(float)
    labels = [str(x) for x in data.index]
    min_v, max_v = values.min(), values.max()
    if min_v == max_v:
        max_v += 1
    chart_w = width - margin_l - margin_r
    chart_h = height - margin_t - margin_b
    draw.line((margin_l, margin_t, margin_l, margin_t + chart_h), fill=(120, 130, 145), width=2)
    draw.line((margin_l, margin_t + chart_h, width - margin_r, margin_t + chart_h), fill=(120, 130, 145), width=2)
    pts = []
    for i, value in enumerate(values):
        x = margin_l + int(i * chart_w / max(len(values) - 1, 1))
        y = margin_t + chart_h - int((value - min_v) / (max_v - min_v) * (chart_h - 30))
        pts.append((x, y))
    if len(pts) > 1:
        draw.line(pts, fill=(216, 77, 68), width=4)
    for i, (x, y) in enumerate(pts):
        draw.ellipse((x - 5, y - 5, x + 5, y + 5), fill=(216, 77, 68))
        if i % 2 == 0 or len(pts) <= 10:
            draw.text((x - 18, margin_t + chart_h + 12), labels[i], fill=(30, 40, 55), font=font)
    img.save(path)


def draw_histogram(values: pd.Series, title: str, path: Path) -> None:
    bins = pd.cut(values, bins=[0, 0.08, 0.25, 0.5, 1, 2, 3, 4], include_lowest=True)
    counts = bins.value_counts().sort_index()
    labels = ["0-1m", "1-3m", "3-6m", "6-12m", "1-2a", "2-3a", "3-4a"]
    counts.index = labels[: len(counts)]
    draw_bar_chart(counts, title, path, color=(62, 146, 131))


def draw_boxplot(values: pd.Series, title: str, path: Path) -> None:
    width, height = 1000, 420
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    font = get_font(16)
    title_font = get_font(26)
    draw.text((70, 25), title, fill=(30, 40, 55), font=title_font)
    q1, med, q3 = values.quantile([0.25, 0.5, 0.75])
    vmin, vmax = values.min(), values.max()
    def sx(v: float) -> int:
        return 90 + int((v - vmin) / (vmax - vmin) * 820) if vmax > vmin else 500
    y = 220
    draw.line((sx(vmin), y, sx(vmax), y), fill=(80, 90, 105), width=3)
    draw.rectangle((sx(q1), y - 55, sx(q3), y + 55), outline=(47, 103, 246), width=4, fill=(229, 238, 255))
    draw.line((sx(med), y - 60, sx(med), y + 60), fill=(216, 77, 68), width=4)
    for v, label in [(vmin, "min"), (q1, "Q1"), (med, "mediana"), (q3, "Q3"), (vmax, "max")]:
        x = sx(float(v))
        draw.line((x, y + 70, x, y + 82), fill=(80, 90, 105), width=2)
        draw.text((x - 35, y + 90), f"{label}\n{v:.2f}", fill=(30, 40, 55), font=font)
    img.save(path)


def draw_grouped_boxplot(final: pd.DataFrame, title: str, path: Path) -> None:
    width, height = 1400, 880
    margin_l, margin_r, margin_t, margin_b = 430, 80, 120, 100
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    font = get_font(16)
    small = get_font(13)
    title_font = get_font(28)
    draw.text((70, 30), title, fill=(30, 40, 55), font=title_font)
    draw.text(
        (70, 68),
        "Idade em meses, por grupo de causa. Caixa = Q1 a Q3; linha vermelha = mediana; pontos = outliers do grupo.",
        fill=(80, 90, 105),
        font=small,
    )

    groups = final["grupo_causa"].value_counts().head(7).index.tolist()
    chart_w = width - margin_l - margin_r
    chart_h = height - margin_t - margin_b
    x_min, x_max = 0, 48

    def sx(value: float) -> int:
        value = max(x_min, min(x_max, float(value)))
        return margin_l + int((value - x_min) / (x_max - x_min) * chart_w)

    axis_y = margin_t + chart_h + 20
    for tick in [0, 6, 12, 24, 36, 48]:
        x = sx(tick)
        draw.line((x, margin_t - 5, x, margin_t + chart_h), fill=(235, 239, 245), width=1)
        draw.line((x, axis_y - 6, x, axis_y + 6), fill=(120, 130, 145), width=2)
        draw.text((x - 12, axis_y + 12), str(tick), fill=(30, 40, 55), font=small)
    draw.line((margin_l, axis_y, width - margin_r, axis_y), fill=(120, 130, 145), width=2)
    draw.text((margin_l + chart_w // 2 - 55, axis_y + 42), "idade em meses", fill=(30, 40, 55), font=font)

    row_gap = chart_h / max(len(groups), 1)
    box_h = 34
    for i, group in enumerate(groups):
        values = final.loc[final["grupo_causa"].eq(group), "idade_meses"].dropna().astype(float).to_numpy()
        if values.size == 0:
            continue
        q1, med, q3 = np.quantile(values, [0.25, 0.5, 0.75])
        iqr = q3 - q1
        lower_fence = q1 - 1.5 * iqr
        upper_fence = q3 + 1.5 * iqr
        inlier_values = values[(values >= lower_fence) & (values <= upper_fence)]
        whisker_low = float(inlier_values.min()) if inlier_values.size else float(values.min())
        whisker_high = float(inlier_values.max()) if inlier_values.size else float(values.max())
        outliers = np.sort(values[(values < lower_fence) | (values > upper_fence)])

        y = int(margin_t + row_gap * i + row_gap / 2)
        label = group if len(group) <= 36 else group[:33] + "..."
        draw.text((70, y - 10), f"{label} (n={values.size})", fill=(30, 40, 55), font=font)
        draw.line((sx(whisker_low), y, sx(whisker_high), y), fill=(80, 90, 105), width=3)
        draw.line((sx(whisker_low), y - 16, sx(whisker_low), y + 16), fill=(80, 90, 105), width=3)
        draw.line((sx(whisker_high), y - 16, sx(whisker_high), y + 16), fill=(80, 90, 105), width=3)
        draw.rectangle((sx(q1), y - box_h // 2, sx(q3), y + box_h // 2), fill=(229, 238, 255), outline=(47, 103, 246), width=3)
        draw.line((sx(med), y - box_h // 2 - 7, sx(med), y + box_h // 2 + 7), fill=(216, 77, 68), width=4)

        if outliers.size:
            shown = outliers
            if outliers.size > 35:
                idx = np.linspace(0, outliers.size - 1, 35).astype(int)
                shown = outliers[idx]
            for j, value in enumerate(shown):
                x = sx(float(value))
                jitter = ((j % 5) - 2) * 3
                draw.ellipse((x - 4, y + jitter - 4, x + 4, y + jitter + 4), fill=(216, 77, 68))
            if outliers.size > shown.size:
                draw.text((sx(float(shown[-1])) + 8, y + 10), f"+{outliers.size - shown.size}", fill=(216, 77, 68), font=small)

        draw.text((sx(med) + 6, y - 35), f"med. {med:.1f}", fill=(80, 90, 105), font=small)

    legend_y = height - 42
    draw.rectangle((70, legend_y - 10, 100, legend_y + 10), fill=(229, 238, 255), outline=(47, 103, 246), width=2)
    draw.text((110, legend_y - 9), "intervalo interquartil", fill=(30, 40, 55), font=small)
    draw.line((310, legend_y - 12, 310, legend_y + 12), fill=(216, 77, 68), width=4)
    draw.text((322, legend_y - 9), "mediana", fill=(30, 40, 55), font=small)
    draw.ellipse((430, legend_y - 5, 440, legend_y + 5), fill=(216, 77, 68))
    draw.text((450, legend_y - 9), "outlier dentro do grupo", fill=(30, 40, 55), font=small)
    img.save(path)


def create_graphs(final: pd.DataFrame, tables: dict) -> list[Path]:
    old_boxplot = GRAPH_DIR / "boxplot_idade.png"
    if old_boxplot.exists():
        old_boxplot.unlink()

    paths = []
    by_year = tables["agg_year"].set_index("ano")["total_obitos"]
    p = GRAPH_DIR / "linha_obitos_por_ano.png"
    draw_line_chart(by_year, "Óbitos por ano", p)
    paths.append(p)

    p = GRAPH_DIR / "barras_obitos_por_ano.png"
    draw_bar_chart(by_year, "Distribuição anual de óbitos", p)
    paths.append(p)

    cause = final["grupo_causa"].value_counts().head(8)
    p = GRAPH_DIR / "barras_grupo_causa.png"
    draw_horizontal_bar_chart(cause, "Principais grupos de causa básica", p)
    paths.append(p)

    p = GRAPH_DIR / "histograma_idade.png"
    draw_histogram(final["idade"], "Distribuição da idade no óbito", p)
    paths.append(p)

    p = GRAPH_DIR / "cruzamento_causa_faixa_etaria.png"
    draw_stacked_horizontal_chart(tables["cause_age_counts"], "Grupo de causa por faixa etária", p)
    paths.append(p)

    external_rates = tables["external_by_age"].set_index("faixa_etaria_bi")["percentual_causas_externas"]
    p = GRAPH_DIR / "percentual_causa_externa_por_faixa.png"
    draw_percent_bar_chart(external_rates, "Peso das causas externas por faixa etária", p)
    paths.append(p)

    p = GRAPH_DIR / "boxplot_idade_por_grupo_causa.png"
    draw_grouped_boxplot(final, "Boxplot da idade por grupo de causa", p)
    paths.append(p)
    return paths


def write_data_catalog(final: pd.DataFrame, meta: dict, tables: dict) -> pd.DataFrame:
    rows = [
        ("ano", "Ano do óbito registrado no SIM.", "Inteiro", "2021", "Base original", "Conversão para inteiro.", "Análise temporal"),
        ("sigla_uf", "Unidade da Federação da residência.", "Texto", "PR", "Base original", "Padronização em maiúsculas.", "Dimensão geográfica"),
        ("id_municipio_residencia", "Código IBGE do município de residência.", "Inteiro", "4108304", "Base original", "Conversão para inteiro.", "Dimensão geográfica"),
        ("causa_basica", "Código CID-10 da causa básica do óbito.", "Texto", "P369", "Base original", "Remoção de espaços e padronização em maiúsculas.", "Dimensão clínica"),
        ("grupo_causa", "Agrupamento analítico da causa básica.", "Texto", "Afecções perinatais", "Criada", "Derivação pelo primeiro caractere do CID-10.", "Segmentação de causas"),
        ("classificacao_evitabilidade", "Classificação simplificada para leitura de evitabilidade.", "Texto", "Potencialmente evitável por atenção à saúde", "Criada", "Regra analítica baseada no grupo de causa.", "Priorização de análise"),
        ("idade", "Idade no óbito em anos.", "Numérico", "0.25", "Base original", "Conversão para número e validação de 0 a 4 anos.", "Medida"),
        ("idade_meses", "Idade convertida para meses.", "Numérico", "3.00", "Criada", "Idade em anos multiplicada por 12.", "Medida interpretável"),
        ("faixa_etaria", "Faixa etária discretizada.", "Categórico", "Pós-neonatal", "Criada", "Discretização por regras de saúde infantil.", "Segmentação"),
        ("faixa_etaria_bi", "Faixa etária consolidada para painéis de BI.", "Categórico", "Menor de 1 ano", "Criada", "Agrupamento em menor de 1 ano e 1 a 4 anos.", "Filtro de dashboard"),
        ("sexo", "Código do sexo informado no registro.", "Inteiro", "1", "Base original", "Conversão para inteiro.", "Dimensão categórica"),
        ("sexo_desc", "Descrição do sexo.", "Categórico", "Masculino", "Criada", "Mapeamento dos códigos de sexo.", "Dimensão categórica"),
        ("raca_cor", "Código de raça/cor.", "Inteiro", "4", "Base original", "Valores ausentes preenchidos com código 9.", "Dimensão sociodemográfica"),
        ("raca_cor_desc", "Descrição de raça/cor.", "Categórico", "Parda", "Criada", "Mapeamento dos códigos de raça/cor.", "Dimensão sociodemográfica"),
        ("local_ocorrencia", "Código do local de ocorrência do óbito.", "Inteiro", "1", "Base original", "Conversão para inteiro.", "Dimensão operacional"),
        ("local_ocorrencia_desc", "Descrição do local de ocorrência.", "Categórico", "Hospital", "Criada", "Mapeamento dos códigos de local.", "Dimensão operacional"),
        ("causa_externa", "Indica se a causa pertence ao grupo de causas externas.", "Booleano", "False", "Criada", "Grupo de causa igual a causas externas.", "Filtro analítico"),
        ("obito_hospitalar", "Indica se o óbito ocorreu em hospital.", "Booleano", "True", "Criada", "Local de ocorrência igual a hospital.", "Indicador"),
        ("idade_outlier_iqr", "Indica outlier estatístico da idade pelo método IQR.", "Booleano", "False", "Criada", "Cálculo por intervalo interquartil.", "Controle de extremos"),
        ("idade_normalizada_minmax", "Idade normalizada entre 0 e 1.", "Numérico", "0.0625", "Criada", "Min-Max Scaling manual.", "Data Mining"),
        ("idade_padronizada_zscore", "Idade padronizada em z-score.", "Numérico", "-0.42", "Criada", "Padronização pela média e desvio padrão.", "Data Mining"),
    ]
    catalog = pd.DataFrame(
        rows,
        columns=[
            "Coluna",
            "Descrição",
            "Tipo",
            "Exemplo",
            "Origem",
            "Tratamento aplicado",
            "Uso",
        ],
    )

    catalog.to_csv(DOC_DIR / "catalogo_dados.csv", sep=";", index=False, encoding="utf-8-sig")

    wb = Workbook()
    ws = wb.active
    ws.title = "Catalogo_Dados"
    ws.append(list(catalog.columns))
    for _, row in catalog.iterrows():
        ws.append(row.tolist())

    ws2 = wb.create_sheet("Resumo")
    ws2.append(["Indicador", "Valor"])
    for key, value in meta["resumo"].items():
        ws2.append([key, value])

    ws3 = wb.create_sheet("Agregacao_Ano")
    ws3.append(list(tables["agg_year"].columns))
    for row in tables["agg_year"].itertuples(index=False):
        ws3.append(list(row))

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="2B2B2B", size=12)
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="CFCFCF")
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if cell.row > 1:
                    cell.font = Font(color="2B2B2B", size=11)
        if sheet.title == "Catalogo_Dados":
            widths = [22, 34, 16, 22, 20, 34, 24]
            for col_idx, width in enumerate(widths, start=1):
                sheet.column_dimensions[get_column_letter(col_idx)].width = width
            sheet.row_dimensions[1].height = 48
            for row_idx in range(2, sheet.max_row + 1):
                sheet.row_dimensions[row_idx].height = 62
        else:
            for col_idx, col in enumerate(sheet.columns, start=1):
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 42)
        sheet.freeze_panes = "A2"

    catalog_path = DOC_DIR / "catalogo_dados.xlsx"
    try:
        wb.save(catalog_path)
    except PermissionError:
        wb.save(DOC_DIR / "catalogo_dados_atualizado.xlsx")
    write_catalog_md(catalog, meta)
    return catalog


def write_catalog_md(catalog: pd.DataFrame, meta: dict) -> None:
    lines = [
        "# Catálogo de Dados - PM3",
        "",
        "## Fonte",
        f"- Base: Sistema de Informação sobre Mortalidade (SIM) - Ministério da Saúde/DATASUS.",
        f"- Link: {SOURCE_URL}",
        f"- Data de acesso: {ACCESS_DATE}",
        f"- Descrição: {SOURCE_DESC}",
        "",
        "## Métricas de processamento",
    ]
    for key, value in meta["resumo"].items():
        lines.append(f"- {key}: {value}")
    lines.extend(["", "## Variáveis do dataset final", ""])
    lines.append("| Coluna | Descrição | Tipo | Exemplo | Origem | Tratamento | Uso |")
    lines.append("|---|---|---|---|---|---|---|")
    for _, row in catalog.iterrows():
        lines.append(
            f"| {row['Coluna']} | {row['Descrição']} | {row['Tipo']} | {row['Exemplo']} | "
            f"{row['Origem']} | {row['Tratamento aplicado']} | {row['Uso']} |"
        )
    (DOC_DIR / "catalogo_dados.md").write_text("\n".join(lines), encoding="utf-8")


def make_table(data: pd.DataFrame, max_rows: int = 10, col_widths: list[float] | None = None) -> Table:
    clipped = data.head(max_rows).copy()
    table_data = [list(clipped.columns)] + clipped.astype(str).values.tolist()
    table = Table(table_data, repeatRows=1, colWidths=col_widths)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9E2F3")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7FAFC")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def add_heading(story: list, text: str, styles) -> None:
    story.append(Paragraph(text, styles["Heading2"]))
    story.append(Spacer(1, 0.15 * cm))


def create_pdf_report(raw: pd.DataFrame, final: pd.DataFrame, no_outliers: pd.DataFrame, meta: dict, tables: dict, graph_paths: list[Path]) -> None:
    output = DOC_DIR / "relatorio_final.pdf"
    try:
        with output.open("ab"):
            pass
    except PermissionError:
        output = DOC_DIR / "relatorio_final_atualizado.pdf"
    doc = SimpleDocTemplate(
        str(output),
        pagesize=A4,
        rightMargin=1.7 * cm,
        leftMargin=1.7 * cm,
        topMargin=1.6 * cm,
        bottomMargin=1.6 * cm,
    )
    styles = getSampleStyleSheet()
    styles["Title"].alignment = TA_CENTER
    styles["Title"].fontSize = 18
    styles["Heading2"].fontSize = 12
    styles["Heading2"].textColor = colors.HexColor("#1F4E78")
    body = ParagraphStyle("BodyCustom", parent=styles["BodyText"], fontSize=9, leading=12, spaceAfter=6)
    small = ParagraphStyle("Small", parent=body, fontSize=8, leading=10)

    under1_count = int((final["faixa_etaria_bi"] == "Menor de 1 ano").sum())
    age_1_4_count = int((final["faixa_etaria_bi"] == "1 a 4 anos").sum())
    under1_pct = under1_count / len(final) * 100
    hospital_pct = final["obito_hospitalar"].mean() * 100
    external_rates = tables["external_by_age"].set_index("faixa_etaria_bi")["percentual_causas_externas"]
    external_under1_pct = float(external_rates.get("Menor de 1 ano", 0))
    external_1_4_pct = float(external_rates.get("1 a 4 anos", 0))
    perinatal_row = tables["cause_age_counts"].set_index("grupo_causa").loc["Afecções perinatais"]
    perinatal_total = int(perinatal_row["total_obitos"])
    perinatal_under1 = int(perinatal_row.get("Menor de 1 ano", 0))
    undefined_home_pct = float(
        tables["local_summary"]
        .set_index("grupo_causa")
        .loc["Sintomas e causas mal definidas", "percentual_domicilio"]
    )
    cause_age_display = tables["cause_age_counts"][
        ["grupo_causa", "Menor de 1 ano", "1 a 4 anos", "total_obitos"]
    ].head(8)
    external_display = tables["external_by_age"][
        ["faixa_etaria_bi", "total_obitos", "causas_externas", "percentual_causas_externas"]
    ]
    local_display = tables["local_summary"][
        ["grupo_causa", "total_obitos", "percentual_hospitalar", "percentual_domicilio"]
    ].head(8)
    q1, q3 = final["idade"].quantile([0.25, 0.75])
    iqr = q3 - q1
    outlier_summary = pd.DataFrame(
        [
            ["Q1 idade", round(float(q1), 3)],
            ["Q3 idade", round(float(q3), 3)],
            ["IQR", round(float(iqr), 3)],
            ["Limite superior", meta["resumo"]["limite_iqr_idade_superior"]],
            ["Registros sinalizados", meta["resumo"]["outliers_idade_iqr"]],
            ["Decisão", "Manter no dataset final e criar versão alternativa sem outliers"],
        ],
        columns=["Item", "Valor"],
    )

    story = [
        Paragraph("Projeto Mensal 3 - Tratamento de Dados", styles["Title"]),
        Paragraph("Base SIM/DATASUS: óbitos de crianças de 0 a 4 anos em Foz do Iguaçu (PR)", styles["Heading3"]),
        Spacer(1, 0.25 * cm),
        Paragraph(f"Fonte oficial: {SOURCE_URL}. Data de acesso: {ACCESS_DATE}.", small),
        Spacer(1, 0.4 * cm),
    ]

    add_heading(story, "1. Introdução", styles)
    story.append(Paragraph(
        "O projeto utiliza uma base real do Sistema de Informação sobre Mortalidade (SIM), "
        "mantido pelo Ministério da Saúde/DATASUS. A base bruta contém registros de óbitos "
        "infantis, com ano, município de residência, causa básica, idade, sexo, raça/cor e "
        "local de ocorrência. O foco do trabalho é preparar os dados antes de qualquer "
        "dashboard ou mineração, garantindo qualidade, padronização e documentação.",
        body,
    ))

    add_heading(story, "2. Tema escolhido", styles)
    story.append(Paragraph(
        "O tema escolhido foi mortalidade de crianças de 0 a 4 anos em Foz do Iguaçu (PR). "
        "A base permite analisar padrões por ano, idade, causa básica, sexo, raça/cor e local "
        "de ocorrência, o que torna o conjunto adequado para tratamento de dados e análise "
        "exploratória.",
        body,
    ))

    add_heading(story, "3. Fonte dos dados", styles)
    story.append(Paragraph(
        f"A fonte é o Sistema de Informação sobre Mortalidade (SIM), disponibilizado pelo "
        f"Ministério da Saúde/DATASUS. Link de referência: {SOURCE_URL}. Data de acesso: "
        f"{ACCESS_DATE}. A base original foi preservada em dados_brutos/dadosSIM2005-.csv.",
        body,
    ))

    add_heading(story, "4. Objetivo da análise", styles)
    story.append(Paragraph(
        "O objetivo é transformar uma base bruta em um dataset final confiável e documentado, "
        "com duplicidades removidas, valores ausentes tratados, categorias padronizadas, "
        "outliers analisados, novas variáveis criadas e evidências exploratórias suficientes "
        "para apoiar BI e Data Mining.",
        body,
    ))

    add_heading(story, "5. Descrição da base original", styles)
    story.append(Paragraph(
        f"A base bruta possui {meta['resumo']['linhas_base_original']} registros e "
        f"{meta['resumo']['colunas_base_original']} colunas. As principais colunas originais "
        "são ano, sigla_uf, id_municipio_residencia, causa_basica, idade, sexo, raca_cor, "
        "escolaridade e local_ocorrencia. Há variáveis numéricas, categóricas codificadas e "
        "campos com valores ausentes, o que atende ao objetivo de tratamento do projeto.",
        body,
    ))

    add_heading(story, "6. Relação da base com BI, Big Data Analytics e Data Mining", styles)
    story.append(Paragraph(
        "Em Business Intelligence, a base pode alimentar dashboards com quantidade de óbitos "
        "por ano, faixa etária, causa básica, local de ocorrência, sexo e raça/cor. Em Big Data "
        "Analytics, o mesmo fluxo poderia ser ampliado para séries históricas nacionais do SIM, "
        "com alto volume, variedade de variáveis e atualização recorrente. Em Data Mining, as "
        "variáveis tratadas permitem agrupamento de perfis, descoberta de padrões por causa, "
        "classificação de registros por risco ou investigação de associações entre causa, idade "
        "e local de ocorrência.",
        body,
    ))

    add_heading(story, "6.1 Modelagem inicial dos dados", styles)
    model_rows = pd.DataFrame(
        [
            ["Linhas na base bruta", meta["resumo"]["linhas_base_original"]],
            ["Colunas na base bruta", meta["resumo"]["colunas_base_original"]],
            ["Linhas no dataset final", meta["resumo"]["linhas_apos_tratamento"]],
            ["Colunas no dataset final", meta["resumo"]["colunas_apos_tratamento"]],
            ["Dimensões para BI", "ano, sexo_desc, raca_cor_desc, local_ocorrencia_desc, grupo_causa, faixa_etaria"],
            ["Medidas para BI", "total de óbitos, idade, idade_meses, indicadores booleanos"],
        ],
        columns=["Item", "Descrição"],
    )
    story.append(make_table(model_rows, max_rows=10, col_widths=[5 * cm, 11 * cm]))
    story.append(Spacer(1, 0.25 * cm))

    add_heading(story, "7. Diagnóstico da qualidade dos dados", styles)
    story.append(Paragraph(
        f"Foram identificadas {meta['resumo']['duplicatas_removidas']} linhas duplicadas, "
        f"{meta['resumo']['faltantes_antes']} valores ausentes na base original e "
        f"{meta['resumo']['escolaridade_faltante_pct']}% de ausência em escolaridade. "
        "A coluna escolaridade foi removida por baixa completude; raça/cor ausente foi "
        "padronizada como Ignorada; categorias codificadas receberam descrições analíticas.",
        body,
    ))
    story.append(make_table(tables["quality"], max_rows=8, col_widths=[3.5 * cm, 3.8 * cm, 2 * cm, 7 * cm]))

    add_heading(story, "8. Análise Exploratória de Dados", styles)
    story.append(Paragraph(
        f"A análise mostrou que {under1_count} registros ({under1_pct:.1f}%) são de menores "
        f"de um ano, contra {age_1_4_count} registros de 1 a 4 anos. Também há predominância "
        f"de ocorrências hospitalares ({hospital_pct:.1f}% do total). A leitura temporal e a "
        "distribuição por causa ajudam a separar o que é volume geral do que é prioridade "
        "específica por faixa etária.",
        body,
    ))
    for p in graph_paths[:4]:
        story.append(RLImage(str(p), width=16 * cm, height=9 * cm))
        story.append(Spacer(1, 0.2 * cm))

    add_heading(story, "8.1 Cruzamentos relevantes para a análise", styles)
    story.append(Paragraph(
        f"O cruzamento entre causa e faixa etária deixa a conclusão mais forte: afecções "
        f"perinatais somam {perinatal_total} registros e todos os {perinatal_under1} com faixa "
        f"informada estão em menores de um ano. Já causas externas representam "
        f"{external_under1_pct:.1f}% dos óbitos em menores de um ano, mas sobem para "
        f"{external_1_4_pct:.1f}% entre 1 e 4 anos. Isso indica dois focos diferentes: cuidado "
        "materno-infantil e neonatal para menores de um ano, e prevenção de acidentes/violências "
        "para crianças de 1 a 4 anos.",
        body,
    ))
    story.append(make_table(cause_age_display, max_rows=8, col_widths=[6.2 * cm, 3.2 * cm, 3 * cm, 3 * cm]))
    story.append(Spacer(1, 0.2 * cm))
    story.append(make_table(external_display, max_rows=4, col_widths=[4.5 * cm, 3 * cm, 3 * cm, 4 * cm]))
    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph(
        f"No cruzamento com local de ocorrência, sintomas e causas mal definidas chamam atenção: "
        f"{undefined_home_pct:.1f}% desses registros ocorreram em domicílio. Esse ponto não deve "
        "ser lido apenas como volume, mas como sinal de qualidade do dado e necessidade de "
        "investigação mais cuidadosa da causa básica.",
        body,
    ))
    story.append(make_table(local_display, max_rows=8, col_widths=[6.2 * cm, 3 * cm, 3.5 * cm, 3.5 * cm]))
    story.append(Spacer(1, 0.2 * cm))
    for p in graph_paths[4:6]:
        story.append(RLImage(str(p), width=16 * cm, height=8 * cm))
        story.append(Spacer(1, 0.2 * cm))
    story.append(PageBreak())

    add_heading(story, "9. Seleção das variáveis", styles)
    story.append(Paragraph(
        "Foram mantidas variáveis úteis para análise temporal, geográfica, demográfica, clínica "
        "e operacional. A coluna escolaridade foi descartada porque a ausência era muito alta e "
        "poderia induzir interpretações frágeis. No dataset final permaneceram variáveis "
        "originais relevantes e variáveis criadas para apoiar análise, como grupo_causa, "
        "idade_meses, faixa_etaria, causa_externa e indicadores de outlier.",
        body,
    ))

    add_heading(story, "10. Limpeza e pré-processamento", styles)
    story.append(Paragraph(
        "As duplicidades exatas foram removidas, textos foram padronizados, códigos foram "
        "convertidos para inteiros, a idade foi convertida para formato numérico e as categorias "
        "codificadas de sexo, raça/cor e local de ocorrência foram mapeadas para descrições "
        "legíveis. Esse pré-processamento preparou a base para análise exploratória e BI.",
        body,
    ))

    add_heading(story, "11. Tratamento de valores faltantes", styles)
    story.append(Paragraph(
        "Antes do tratamento havia valores ausentes em raça/cor e escolaridade. Depois do "
        "tratamento, o dataset final ficou sem valores ausentes: raça/cor recebeu a categoria "
        "Ignorada e escolaridade foi removida com justificativa técnica.",
        body,
    ))
    story.append(make_table(meta["faltantes"], max_rows=12, col_widths=[6 * cm, 4 * cm, 4 * cm]))

    add_heading(story, "12. Tratamento de outliers", styles)
    story.append(Paragraph(
        f"O método IQR aplicado à idade gerou limite superior de "
        f"{meta['resumo']['limite_iqr_idade_superior']} ano e identificou "
        f"{meta['resumo']['outliers_idade_iqr']} registros como extremos estatísticos. "
        "Esses registros não foram tratados como erro automático, pois óbitos entre 1 e 4 anos "
        "são possíveis e relevantes. A decisão principal foi mantê-los no dataset final com a "
        "coluna idade_outlier_iqr e também gerar uma versão alternativa sem outliers.",
        body,
    ))
    story.append(make_table(outlier_summary, max_rows=8, col_widths=[5 * cm, 10 * cm]))
    story.append(Paragraph(
        "Para visualizar outliers sem distorcer a interpretação, foi usado um boxplot da idade "
        "em meses por grupo de causa. Essa versão é mais informativa que o boxplot único, pois "
        "mostra que alguns valores extremos não são erros: eles representam perfis diferentes "
        "de mortalidade, como causas externas e sistema nervoso ocorrendo em idades mais altas.",
        body,
    ))
    boxplot_path = GRAPH_DIR / "boxplot_idade_por_grupo_causa.png"
    if boxplot_path.exists():
        story.append(RLImage(str(boxplot_path), width=16 * cm, height=10 * cm))
        story.append(Spacer(1, 0.2 * cm))

    add_heading(story, "13. Transformações realizadas", styles)
    story.append(Paragraph(
        "Foram criadas descrições para códigos de sexo, raça/cor e local de ocorrência; "
        "agrupamento de causa por capítulo simplificado do CID-10; idade em meses; indicadores "
        "de causa externa e óbito hospitalar; faixa etária para BI; e classificação simplificada "
        "de evitabilidade.",
        body,
    ))

    add_heading(story, "14. Agregações realizadas", styles)
    story.append(Paragraph(
        "A principal agregação resume o total de óbitos por ano. Ela ajuda a identificar anos "
        "com maior volume de registros e serve como base para dashboards temporais. Também "
        "foram geradas agregações por faixa etária, grupo de causa e local de ocorrência para "
        "sustentar os cruzamentos exploratórios.",
        body,
    ))
    story.append(make_table(tables["agg_year"], max_rows=25, col_widths=[5 * cm, 5 * cm]))

    add_heading(story, "15. Normalização e padronização", styles)
    story.append(Paragraph(
        "A variável idade foi normalizada com Min-Max Scaling, variando de 0 a 1, e também "
        "padronizada em z-score. A normalização facilita comparações em escala comum, enquanto "
        "o z-score indica quantos desvios padrão cada registro está distante da média. Essas "
        "duas versões são úteis para técnicas de Data Mining e Machine Learning.",
        body,
    ))

    add_heading(story, "16. Discretização", styles)
    story.append(Paragraph(
        "A variável numérica idade foi transformada em categorias analíticas. A coluna "
        "faixa_etaria separa registros em Neonatal, Pós-neonatal, 1 ano, 2 anos, 3 anos e "
        "4 anos. A coluna faixa_etaria_bi consolida a leitura em Menor de 1 ano e 1 a 4 anos, "
        "facilitando a construção de filtros e gráficos em BI.",
        body,
    ))

    add_heading(story, "17. Feature Engineering", styles)
    story.append(Paragraph(
        "As principais variáveis criadas foram grupo_causa, classificacao_evitabilidade, "
        "idade_meses, faixa_etaria, faixa_etaria_bi, causa_externa, obito_hospitalar, "
        "idade_outlier_iqr, idade_normalizada_minmax e idade_padronizada_zscore. Essas "
        "features aumentam a utilidade analítica do dataset sem alterar a base original.",
        body,
    ))

    add_heading(story, "18. Descrição do dataset final", styles)
    story.append(Paragraph(
        "O dataset final está limpo, sem duplicidades exatas, sem valores faltantes, com nomes "
        f"padronizados, variáveis originais relevantes e variáveis criadas. Ele possui "
        f"{meta['resumo']['linhas_apos_tratamento']} registros e "
        f"{meta['resumo']['colunas_apos_tratamento']} colunas, pronto para uso em BI, "
        "Data Mining e análise exploratória.",
        body,
    ))

    add_heading(story, "19. Catálogo de dados", styles)
    story.append(Paragraph(
        "O catálogo de dados foi gerado em formato XLSX, CSV e Markdown. Ele documenta, para "
        "cada coluna do dataset final, o nome da coluna, descrição, tipo, exemplo, origem, "
        "tratamento aplicado e uso esperado na análise.",
        body,
    ))

    add_heading(story, "20. Organização DataOps", styles)
    story.append(Paragraph(
        "A organização do projeto preserva a base bruta, separa dados tratados, mantém o "
        "notebook/script reprodutível e concentra a documentação na pasta documentacao. Também "
        "foi gerada uma pasta final chamada PM3_Tratamento_Dados com os arquivos necessários "
        "para entrega e reprodução do processamento.",
        body,
    ))
    dataops_rows = pd.DataFrame(
        [
            ["dados_brutos/", "Base original preservada."],
            ["dados_tratados/", "Dataset final tratado e versão sem outliers estatísticos."],
            ["notebooks/", "Notebook tratamento_dados_pm3.ipynb."],
            ["documentacao/", "Relatório final, catálogo, tabelas de evidências e gráficos."],
            ["projMens3.py", "Script Python usado para reproduzir todo o processo."],
            ["README.md", "Instruções de execução e descrição dos entregáveis."],
        ],
        columns=["Pasta/arquivo", "Finalidade"],
    )
    story.append(make_table(dataops_rows, max_rows=10, col_widths=[5 * cm, 10 * cm]))

    add_heading(story, "21. Conclusão", styles)
    story.append(Paragraph(
        "A conclusão analítica principal é que a base não deve ser lida apenas pelo total de "
        "óbitos. O cruzamento das variáveis mostra perfis diferentes: menores de um ano estão "
        "mais associados a causas perinatais e malformações congênitas, enquanto a faixa de "
        "1 a 4 anos apresenta peso proporcional muito maior de causas externas. Assim, um "
        "dashboard útil deveria separar esses grupos em vez de apresentar somente um ranking "
        "geral de causas.",
        body,
    ))
    story.append(Paragraph(
        "Outro achado relevante é a relação entre local de ocorrência e qualidade da causa "
        "registrada: causas mal definidas têm participação domiciliar alta em relação aos "
        "principais grupos. Isso sugere que parte do uso da base em BI deve monitorar qualidade "
        "do preenchimento, não apenas indicadores finais de mortalidade.",
        body,
    ))

    add_heading(story, "22. Limitações", styles)
    story.append(Paragraph(
        "A base está restrita a um município e a classificação de evitabilidade usada no projeto "
        "é simplificada para fins didáticos, não substituindo protocolos oficiais de investigação "
        "epidemiológica. Além disso, a análise trabalha com contagens absolutas e não com taxas "
        "populacionais.",
        body,
    ))

    add_heading(story, "23. Próximos passos", styles)
    story.append(Paragraph(
        "Os próximos passos recomendados são cruzar o SIM com dados populacionais, calcular "
        "taxas por nascidos vivos, ampliar a comparação com outros municípios e construir um "
        "dashboard temporal com filtros por faixa etária, grupo de causa, local de ocorrência, "
        "sexo e raça/cor.",
        body,
    ))

    doc.build(story)


def write_notebook() -> None:
    cells = [
        {
            "cell_type": "markdown",
            "metadata": {},
            "source": [
                "# PM3 - Tratamento de Dados\n",
                "\n",
                "Notebook reprodutível do Projeto Mensal 3 usando dados reais do SIM/DATASUS.\n",
            ],
        },
        {
            "cell_type": "markdown",
            "metadata": {},
            "source": [
                "## Fonte dos dados\n",
                f"- Sistema de Informação sobre Mortalidade (SIM): {SOURCE_URL}\n",
                f"- Data de acesso: {ACCESS_DATE}\n",
            ],
        },
        {
            "cell_type": "code",
            "execution_count": None,
            "metadata": {},
            "outputs": [],
            "source": [
                "from pathlib import Path\n",
                "import pandas as pd\n",
                "import numpy as np\n",
                "\n",
                "cwd = Path.cwd().resolve()\n",
                "if (cwd / 'dados_brutos').exists():\n",
                "    base_dir = cwd\n",
                "elif cwd.name == 'notebooks' and (cwd.parent / 'dados_brutos').exists():\n",
                "    base_dir = cwd.parent\n",
                "else:\n",
                "    base_dir = Path('..').resolve()\n",
                "raw_path = base_dir / 'dados_brutos' / 'dadosSIM2005-.csv'\n",
                "df = pd.read_csv(raw_path)\n",
                "df.shape, df.head()\n",
            ],
        },
        {
            "cell_type": "code",
            "execution_count": None,
            "metadata": {},
            "outputs": [],
            "source": [
                "# Diagnóstico inicial de qualidade\n",
                "diagnostico = pd.DataFrame({\n",
                "    'tipo': ['linhas', 'colunas', 'duplicatas', 'faltantes'],\n",
                "    'valor': [len(df), df.shape[1], df.duplicated().sum(), df.isna().sum().sum()]\n",
                "})\n",
                "diagnostico\n",
            ],
        },
        {
            "cell_type": "code",
            "execution_count": None,
            "metadata": {},
            "outputs": [],
            "source": [
                "# Execução do pipeline completo\n",
                "import sys\n",
                "sys.path.append(str(base_dir))\n",
                "import projMens3\n",
                "projMens3.main()\n",
            ],
        },
        {
            "cell_type": "code",
            "execution_count": None,
            "metadata": {},
            "outputs": [],
            "source": [
                "final = pd.read_csv(base_dir / 'dados_tratados' / 'dataset_final_tratado.csv', sep=';')\n",
                "final.head(), final.shape\n",
            ],
        },
        {
            "cell_type": "code",
            "execution_count": None,
            "metadata": {},
            "outputs": [],
            "source": [
                "# Estatísticas descritivas e frequências categóricas\n",
                "display(final[['idade', 'idade_meses', 'idade_normalizada_minmax', 'idade_padronizada_zscore']].describe())\n",
                "display(final['grupo_causa'].value_counts().head(10))\n",
                "display(final.groupby('ano').size().rename('total_obitos').reset_index())\n",
            ],
        },
        {
            "cell_type": "markdown",
            "metadata": {},
            "source": [
                "## Cruzamentos para conclusão analítica\n",
                "\n",
                "Aqui a análise deixa de olhar apenas frequências isoladas e cruza causa, faixa etária e local de ocorrência.\n",
            ],
        },
        {
            "cell_type": "code",
            "execution_count": None,
            "metadata": {},
            "outputs": [],
            "source": [
                "cruzamento_causa_faixa = pd.crosstab(final['grupo_causa'], final['faixa_etaria_bi'])\n",
                "cruzamento_causa_faixa['total_obitos'] = cruzamento_causa_faixa.sum(axis=1)\n",
                "display(cruzamento_causa_faixa.sort_values('total_obitos', ascending=False).head(10))\n",
                "\n",
                "causa_externa_por_faixa = final.groupby('faixa_etaria_bi')['causa_externa'].agg(['count', 'sum'])\n",
                "causa_externa_por_faixa['percentual_causas_externas'] = (causa_externa_por_faixa['sum'] / causa_externa_por_faixa['count'] * 100).round(1)\n",
                "display(causa_externa_por_faixa)\n",
            ],
        },
        {
            "cell_type": "markdown",
            "metadata": {},
            "source": [
                "Conclusão: os menores de 1 ano concentram causas perinatais e malformações congênitas. Já entre 1 e 4 anos, o peso proporcional das causas externas aumenta bastante, o que muda a interpretação para BI e políticas de prevenção.\n",
            ],
        },
        {
            "cell_type": "markdown",
            "metadata": {},
            "source": [
                "## Onde os gráficos são gerados\n",
                "\n",
                "Ao executar `projMens3.main()`, os gráficos são salvos como imagens PNG em `documentacao/graficos/` e também entram no `documentacao/relatorio_final.pdf`.\n",
            ],
        },
        {
            "cell_type": "code",
            "execution_count": None,
            "metadata": {},
            "outputs": [],
            "source": [
                "from IPython.display import Image, display\n",
                "\n",
                "graph_dir = base_dir / 'documentacao' / 'graficos'\n",
                "graficos = sorted(graph_dir.glob('*.png'))\n",
                "print(f'{len(graficos)} gráficos gerados em: {graph_dir}')\n",
                "for grafico in graficos:\n",
                "    print(grafico.name)\n",
                "    display(Image(filename=str(grafico)))\n",
            ],
        },
    ]
    nb = {
        "cells": cells,
        "metadata": {
            "kernelspec": {
                "display_name": "Python 3",
                "language": "python",
                "name": "python3",
            },
            "language_info": {
                "name": "python",
                "pygments_lexer": "ipython3",
            },
        },
        "nbformat": 4,
        "nbformat_minor": 5,
    }
    (NOTEBOOK_DIR / "tratamento_dados_pm3.ipynb").write_text(json.dumps(nb, ensure_ascii=False, indent=2), encoding="utf-8")


def latest_existing(*paths: Path) -> Path:
    existing = [path for path in paths if path.exists()]
    if not existing:
        raise FileNotFoundError("Nenhum dos arquivos esperados foi encontrado.")
    return max(existing, key=lambda path: path.stat().st_mtime)


def copy_file(src: Path, dst: Path) -> None:
    if not src.exists():
        return
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)


def copy_tree_contents(src_dir: Path, dst_dir: Path) -> None:
    if not src_dir.exists():
        return
    for src in src_dir.rglob("*"):
        if src.is_file():
            rel = src.relative_to(src_dir)
            copy_file(src, dst_dir / rel)


def create_delivery_package() -> None:
    if BASE_DIR.name == "PM3_Tratamento_Dados":
        return

    DELIVERY_DIR.mkdir(exist_ok=True)
    for folder in ["dados_brutos", "dados_tratados", "notebooks", "documentacao"]:
        copy_tree_contents(BASE_DIR / folder, DELIVERY_DIR / folder)

    copy_file(RAW_PATH, DELIVERY_DIR / "dados_brutos" / "base_original.csv")
    copy_file(BASE_DIR / "projMens3.py", DELIVERY_DIR / "projMens3.py")
    copy_file(BASE_DIR / "executar_projeto.bat", DELIVERY_DIR / "executar_projeto.bat")
    copy_file(BASE_DIR / "README.md", DELIVERY_DIR / "README.md")
    copy_file(BASE_DIR / "requirements.txt", DELIVERY_DIR / "requirements.txt")

    latest_catalog = latest_existing(
        DOC_DIR / "catalogo_dados.xlsx",
        DOC_DIR / "catalogo_dados_atualizado.xlsx",
    )
    copy_file(latest_catalog, DELIVERY_DIR / "documentacao" / "catalogo_dados.xlsx")

    latest_report = latest_existing(
        DOC_DIR / "relatorio_final.pdf",
        DOC_DIR / "relatorio_final_atualizado.pdf",
    )
    copy_file(latest_report, DELIVERY_DIR / "documentacao" / "relatorio_final.pdf")

    for extra_name in ["catalogo_dados_atualizado.xlsx", "relatorio_final_atualizado.pdf"]:
        extra_path = DELIVERY_DIR / "documentacao" / extra_name
        if extra_path.exists():
            extra_path.unlink()


def validate_outputs() -> None:
    final = pd.read_csv(TREATED_DIR / "dataset_final_tratado.csv", sep=";")
    assert final.shape[0] > 300, "Dataset final precisa ter mais de 300 registros."
    assert final.shape[1] >= 6, "Dataset final precisa ter pelo menos 6 colunas."
    assert final.duplicated().sum() == 0, "Dataset final contém duplicidades exatas."
    assert final.isna().sum().sum() == 0, "Dataset final contém valores ausentes."
    required = [
        DOC_DIR / "relatorio_final.pdf",
        DOC_DIR / "catalogo_dados.xlsx",
        DOC_DIR / "problemas_qualidade.csv",
        NOTEBOOK_DIR / "tratamento_dados_pm3.ipynb",
    ]
    for path in required:
        assert path.exists() and path.stat().st_size > 0, f"Arquivo ausente ou vazio: {path}"

    catalog_path = latest_existing(DOC_DIR / "catalogo_dados.xlsx", DOC_DIR / "catalogo_dados_atualizado.xlsx")
    wb = load_workbook(catalog_path, read_only=True)
    assert "Catalogo_Dados" in wb.sheetnames, "Catálogo XLSX sem aba Catalogo_Dados."
    ws = wb["Catalogo_Dados"]
    expected_headers = ["Coluna", "Descrição", "Tipo", "Exemplo", "Origem", "Tratamento aplicado", "Uso"]
    headers = [ws.cell(1, i).value for i in range(1, 8)]
    assert headers == expected_headers, "Catálogo XLSX fora do formato exigido."
    wb.close()


def main() -> None:
    ensure_dirs()
    raw = read_raw()
    final, no_outliers, meta = treat_data(raw)
    tables = save_tables(raw, final, no_outliers, meta)
    graph_paths = create_graphs(final, tables)
    catalog = write_data_catalog(final, meta, tables)
    create_pdf_report(raw, final, no_outliers, meta, tables, graph_paths)
    write_notebook()
    validate_outputs()
    create_delivery_package()
    print(json.dumps(meta["resumo"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
